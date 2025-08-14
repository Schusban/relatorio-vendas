import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import io
import zipfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
import tempfile
import os
from datetime import datetime
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors


# =============================
# Funções Auxiliares
# =============================

def quebrar_por_vendedor(df: pd.DataFrame, pasta_destino: str) -> list:
    """
    Cria um arquivo ZIP contendo os arquivos informados.

    Parâmetros:
    arquivos (list): Lista de caminhos para os arquivos que serão compactados.

    Retorna:
    io.BytesIO: Objeto de memória contendo o arquivo ZIP pronto para download.
    """
    arquivos_gerados = []
    for vendedor, grupo in df.groupby("Vendedor"):
        caminho_arquivo = os.path.join(pasta_destino, f"{vendedor}.xlsx")
        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:
            grupo.to_excel(writer, index=False)
        arquivos_gerados.append(caminho_arquivo)
    return arquivos_gerados


def criar_planilha_resumo(df: pd.DataFrame, df_resumo: pd.DataFrame, pasta_destino: str) -> str:
    """
    Cria uma planilha Excel contendo um resumo de vendas e abas separadas por vendedor.

    Funcionalidades:
    - Aba "Resumo": total de vendas por vendedor.
    - Aba "Gráficos": insere dois gráficos (barras e pizza).
    - Uma aba para cada vendedor com seus respectivos registros.

    Parâmetros:
        df (pd.DataFrame): DataFrame completo com as vendas detalhadas (colunas: "Vendedor", "Produto", "Vendas"...).
        df_resumo (pd.DataFrame): DataFrame resumido com o total de vendas por vendedor.
        pasta_destino (str): Caminho da pasta onde o arquivo será salvo.

    Retorno:
        str: Caminho completo do arquivo Excel gerado.

    Observações:
        - Os gráficos são gerados em PNG temporários e inseridos na aba "Gráficos".
        - O nome das abas de vendedor é limitado a 31 caracteres (limite do Excel).
    """
    caminho_arquivo = os.path.join(pasta_destino, "planilha_resumo.xlsx")
    wb = Workbook()

    # Aba Resumo
    aba_resumo = wb.active
    aba_resumo.title = "Resumo"
    for r in dataframe_to_rows(df_resumo, index=False, header=True):
        aba_resumo.append(r)

    # Gerar gráficos e salvar imagens temporárias
    grafico_vendedor = gerar_grafico_vendas(df_resumo.rename(columns={"Total de Vendas": "Vendas"}), "vendedor")
    grafico_produto = gerar_grafico_vendas(df, "produto")

    img_vendedor_path = os.path.join(pasta_destino, "grafico_vendedor.png")
    img_produto_path = os.path.join(pasta_destino, "grafico_produto.png")
    grafico_vendedor.savefig(img_vendedor_path, dpi=100, bbox_inches='tight')
    grafico_produto.savefig(img_produto_path, dpi=100, bbox_inches='tight')

    # Aba Gráficos
    aba_graficos = wb.create_sheet("Gráficos")
    aba_graficos.add_image(XLImage(img_vendedor_path), "A1")
    aba_graficos.add_image(XLImage(img_produto_path), "A20")

    # Abas por vendedor
    for vendedor, grupo in df.groupby("Vendedor"):
        aba_vendedor = wb.create_sheet(title=str(vendedor)[:31])
        for r in dataframe_to_rows(grupo, index=False, header=True):
            aba_vendedor.append(r)

    wb.save(caminho_arquivo)
    return caminho_arquivo


def criar_zip(arquivos: list) -> io.BytesIO:
    """
    Cria um arquivo ZIP contendo os arquivos especificados.

    Parâmetros:
        arquivos (list): Lista de caminhos completos dos arquivos a serem incluídos no ZIP.

    Retorno:
        io.BytesIO: Objeto em memória contendo o arquivo ZIP pronto para ser salvo ou transmitido.

    Observações:
        - Os arquivos no ZIP manterão apenas seus nomes originais, não o caminho completo.
    """

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for arquivo in arquivos:
            z.write(arquivo, os.path.basename(arquivo))
    buffer.seek(0)
    return buffer


def gerar_grafico_vendas(df: pd.DataFrame, tipo: str = "vendedor"):
    """
    Gera gráficos de vendas no formato matplotlib.

    Tipos de gráficos:
        - "vendedor": gráfico de barras com total de vendas por vendedor.
        - "produto": gráfico de pizza com proporção de vendas por produto.

    Parâmetros:
        df (pd.DataFrame): DataFrame com dados de vendas (deve conter colunas "Vendedor" ou "Produto" e "Vendas").
        tipo (str): Tipo de gráfico a gerar ("vendedor" ou "produto").

    Retorno:
        matplotlib.figure.Figure: Objeto Figure contendo o gráfico gerado.

    Observações:
        - O gráfico de pizza posiciona porcentagens e rótulos externamente para melhor leitura.
        - As fontes são reduzidas para adequar-se a exportações em tamanhos pequenos.
    """

    fig, ax = plt.subplots(figsize=(5, 3))
    plt.rcParams.update({'font.size': 7})

    if tipo == "vendedor":
        resumo = df.groupby("Vendedor")["Vendas"].sum().sort_values(ascending=False)
        barras_grafico = ax.bar(resumo.index, resumo.values, color="skyblue", edgecolor="black")
        ax.set_title("Vendas por Vendedor", fontsize=9)
        ax.set_xlabel("Vendedor", fontsize=7)
        ax.set_ylabel("Total de Vendas", fontsize=7)
        ax.bar_label(barras_grafico, fmt="%.2f", padding=2, fontsize=5)
        ax.tick_params(axis='x', labelrotation=45, labelsize=5)
        ax.tick_params(axis='y', labelsize=5)

    elif tipo == "produto":
        resumo = df.groupby("Produto")["Vendas"].sum().sort_values(ascending=False)
        
        # Gerar cores únicas usando um colormap
        cmap = plt.get_cmap("tab20")  # tab20 tem 20 cores distintas
        num_produtos = len(resumo)
        cores = [cmap(i / num_produtos) for i in range(num_produtos)]

        wedges, texts, autotexts = ax.pie(
            resumo.values,
            autopct="%.1f%%",
            startangle=90,
            textprops={'fontsize': 5},
            pctdistance=1.15,
            labeldistance=1.3,
            colors=cores
        )
        
        # Fonte das porcentagens
        for autotext in autotexts:
            autotext.set_fontsize(5)
            autotext.set_color('black')

        # Legenda ao lado
        ax.legend(
            wedges,
            resumo.index,
            title="Produtos",
            loc="center left",
            bbox_to_anchor=(1, 0, 0.5, 1),
            fontsize=5,
            title_fontsize=6
        )

        ax.set_title("Proporção de Vendas por Produto", fontsize=9)

    plt.tight_layout()
    return fig



def salvar_relatorio_completo_em_pdf(df_resumo: pd.DataFrame, df: pd.DataFrame, caminho_pdf: str):
    """
    Gera um relatório PDF contendo:
    - Título e data de geração.
    - Tabela resumo de vendas por vendedor.
    - Gráfico de barras (vendas por vendedor).
    - Gráfico de pizza (proporção de vendas por produto).

    Parâmetros:
        df_resumo (pd.DataFrame): DataFrame resumido (colunas: "Vendedor", "Total de Vendas").
        df (pd.DataFrame): DataFrame completo das vendas.
        caminho_pdf (str): Caminho completo para salvar o arquivo PDF.

    Retorno:
        None

    Observações:
        - Os gráficos são salvos temporariamente como PNG e embutidos no PDF.
        - O layout utiliza a biblioteca reportlab com estilos pré-definidos.
    """

    # Criar imagens temporárias dos gráficos
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img_vendedor, \
         tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img_produto:
        
        grafico_vendedor = gerar_grafico_vendas(df_resumo.rename(columns={"Total de Vendas": "Vendas"}), "vendedor")
        grafico_produto = gerar_grafico_vendas(df, "produto")
        grafico_vendedor.savefig(tmp_img_vendedor.name, format="png", dpi=100, bbox_inches='tight')
        grafico_produto.savefig(tmp_img_produto.name, format="png", dpi=100, bbox_inches='tight')

        # Criar documento PDF
        doc = SimpleDocTemplate(caminho_pdf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Título
        story.append(Paragraph("Relatório de Vendas", styles["Title"]))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Tabela resumo
        tabela_dados = [list(df_resumo.columns)] + df_resumo.values.tolist()
        tabela = Table(tabela_dados)
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(Paragraph("Resumo de Vendas por Vendedor:", styles["Heading2"]))
        story.append(tabela)
        story.append(Spacer(1, 20))

        # Gráfico 1 - Barras
        story.append(Paragraph("Vendas por Vendedor", styles["Heading2"]))
        story.append(RLImage(tmp_img_vendedor.name, width=400, height=250))
        story.append(Spacer(1, 20))

        # Gráfico 2 - Pizza
        story.append(Paragraph("Proporção de Vendas por Produto", styles["Heading2"]))
        story.append(RLImage(tmp_img_produto.name, width=400, height=250))

        # Finalizar PDF
        doc.build(story)


# =============================
# Interface Streamlit
# =============================
st.set_page_config(page_title="📊 Relatórios de Vendas", layout="wide")
st.title("📊 Sistema de Relatórios de Vendas")
st.write("Faça upload da planilha Excel e gere arquivos por vendedor, resumo e relatórios.")

st.markdown("""
### 📄 Estrutura da planilha necessária
A planilha deve conter **exatamente** as seguintes colunas:
- **Vendedor** → Nome do vendedor.
- **Produto** → Nome do produto vendido.
- **Vendas** → Valor total da venda (numérico).

Exemplo:
| Vendedor | Produto  | Vendas |
|----------|----------|--------|
| João     | Caneta   | 150.50 |
| Maria    | Caderno  | 320.00 |

Você pode baixar um **modelo pronto** para preencher:
""")

# Criar planilha modelo
df_modelo = pd.DataFrame({
    "Vendedor": ["João", "Maria"],
    "Produto": ["Caneta", "Caderno"],
    "Vendas": [150.50, 320.00]
})
modelo_buffer = io.BytesIO()
df_modelo.to_excel(modelo_buffer, index=False)
modelo_buffer.seek(0)

st.download_button(
    label="📥 Baixar Modelo Excel",
    data=modelo_buffer,
    file_name="modelo_vendas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


arquivo = st.file_uploader("Selecione a planilha Excel", type=["xlsx"])

if arquivo:
    df = pd.read_excel(arquivo)

    # Validação das colunas
    colunas_necessarias = {"Vendedor", "Produto", "Vendas"}
    if not colunas_necessarias.issubset(df.columns):
        st.error(f"O arquivo precisa conter as colunas: {', '.join(colunas_necessarias)}")
    else:
        st.success("Arquivo carregado com sucesso ✅")

        # Resumo
        df_resumo = df.groupby("Vendedor", as_index=False)["Vendas"].sum().rename(columns={"Vendas": "Total de Vendas"})
        st.subheader("📋 Resumo de Vendas por Vendedor")
        st.dataframe(df_resumo)

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Criar arquivos por vendedor
            arquivos_vendedores = quebrar_por_vendedor(df, tmp_dir)

            # Criar planilha resumo (com gráficos e abas por vendedor)
            caminho_resumo = criar_planilha_resumo(df, df_resumo, tmp_dir)

            # Criar ZIP final
            arquivos_zip = arquivos_vendedores + [caminho_resumo]
            zip_buffer = criar_zip(arquivos_zip)

            st.markdown("""
            ### **O que vem no arquivo ZIP?**
            - 📂 **Arquivos por vendedor** → Cada vendedor terá seu próprio arquivo Excel.
            - 📊 **Planilha resumo** → Contém:
                - Aba "Resumo" com total de vendas por vendedor.
                - Aba "Gráficos" com visualizações de vendas.
                - Abas individuais para cada vendedor.
            """)

            st.download_button(
                label="📦 Baixar Arquivos ZIP",
                data=zip_buffer,
                file_name="relatorios_vendas.zip",
                mime="application/zip"
            )

            # PDF para download
            caminho_pdf = os.path.join(tmp_dir, "relatorio_vendas.pdf")
            salvar_relatorio_completo_em_pdf(df_resumo, df, caminho_pdf)

            st.markdown("""
            ### **O que vem no Relatório PDF?**
            - 📝 **Título e data** de geração.
            - 📋 **Tabela resumo** com vendas por vendedor.
            - 📈 **Gráfico de barras** com vendas por vendedor.
            - 🥧 **Gráfico de pizza** com proporção de vendas por produto.
            """)


            with open(caminho_pdf, "rb") as pdf_file:
                st.download_button(
                    label="📑 Baixar Relatório PDF",
                    data=pdf_file,
                    file_name="relatorio_vendas.pdf",
                    mime="application/pdf"
                )

            # Exibir gráficos
            st.subheader("📈 Visualizações")
            col1, col2 = st.columns(2)
            col1.pyplot(gerar_grafico_vendas(df_resumo.rename(columns={"Total de Vendas": "Vendas"}), "vendedor"))
            col2.pyplot(gerar_grafico_vendas(df, "produto"))
